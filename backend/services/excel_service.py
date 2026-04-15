import openpyxl
import re
from datetime import datetime

class ExcelService:
    def __init__(self):
        pass
    
    def process_excel(self, excel_path):
        """
        处理Excel文件并返回结构化数据
        :param excel_path: Excel文件路径
        :return: 按题材分组的股票数据和识别到的日期
        """
        try:
            # 打开Excel文件
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            # 提取日期
            date = self._extract_date(worksheet)
            
            # 提取股票数据
            data = self._extract_stock_data(worksheet)
            
            return {
                'success': True,
                'data': data,
                'date': date.isoformat() if date else None
            }
        except Exception as e:
            return {'error': f'处理Excel文件失败: {str(e)}'}
    
    def _extract_date(self, worksheet):
        """
        从Excel中提取日期
        :param worksheet: Excel工作表
        :return: 日期对象
        """
        # 检查第一行是否包含日期信息
        for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    # 匹配日期模式
                    date_patterns = [
                        r'(\d{4})[-/](0[1-9]|1[0-2])[-/](0[1-9]|[12]\d|3[01])',  # 2024-04-02
                        r'(\d{4})年(0?[1-9]|1[0-2])月(0?[1-9]|[12]\d|3[01])日',  # 2024年4月2日
                        r'\((0[1-9]|1[0-2])[./](0[1-9]|[12]\d|3[01])\)',  # (04.02) 格式
                    ]
                    for pattern in date_patterns:
                        match = re.search(pattern, cell)
                        if match:
                            try:
                                # 尝试解析日期
                                if len(match.groups()) == 3:
                                    # 完整日期格式 (2024-04-02 或 2024年4月2日)
                                    year, month, day = match.groups()
                                    return datetime(int(year), int(month), int(day)).date()
                                elif len(match.groups()) == 2:
                                    # 简短日期格式 (04.02)
                                    month, day = match.groups()
                                    # 使用当前年份
                                    year = datetime.now().year
                                    return datetime(int(year), int(month), int(day)).date()
                            except:
                                pass
        return None
    
    def _extract_stock_data(self, worksheet):
        """
        从Excel中提取股票数据
        :param worksheet: Excel工作表
        :return: 按题材分组的字典
        """
        result = {}
        current_theme = None
        current_stocks = []
        is_header = True
        
        # 遍历所有行
        for row in worksheet.iter_rows(values_only=True):
            # 过滤空行
            if not any(cell for cell in row):
                continue
            
            # 检查是否是表头
            if is_header:
                # 检查是否包含表头关键词
                row_str = [str(cell) if cell is not None else '' for cell in row]
                line = ' '.join(row_str)
                if '板数' in line and '代码' in line and '个股' in line:
                    is_header = False
                continue
            
            # 检查是否是题材行：A列有内容，B列没有内容
            if len(row) > 0 and row[0] and (len(row) <= 1 or not row[1]):
                # 保存上一个题材的数据
                if current_theme and current_stocks:
                    result[current_theme] = current_stocks
                
                # 提取题材名称
                theme_name = str(row[0]).strip()
                # 清理题材名称，移除可能的标记和特殊字符
                theme_name = re.sub(r'[*×+]\d*$', '', theme_name)
                theme_name = re.sub(r'\d+[*×+]$', '', theme_name)
                # 清理题材名称，移除数字和特殊字符，保留中文和英文字母
                theme_name = re.sub(r'[\d*×+​]', '', theme_name)
                theme_name = theme_name.strip()
                
                if theme_name and self._is_valid_theme(theme_name):
                    current_theme = theme_name
                    current_stocks = []
                continue
            
            # 尝试识别股票行
            stock_info = self._extract_stock_line(row)
            if stock_info and current_theme:
                current_stocks.append(stock_info)
        
        # 保存最后一个题材的数据
        if current_theme and current_stocks:
            result[current_theme] = current_stocks
        
        return result
    
    def _extract_theme(self, line):
        """
        提取题材名称
        匹配模式：
        - 医疗医药*7
        - 美伊战争*6
        """
        # 清理行
        line = line.strip()
        
        # 匹配题材名称（中文+可能的数字标记）
        patterns = [
            r'^([\u4e00-\u9fa5]+)[\*×]\d+',  # 匹配 "医疗医药*7" 或 "医疗医药×7"
            r'^([\u4e00-\u9fa5]+)\d+[\*×]',  # 匹配 "AI硬件7*" 或 "AI硬件7×"
            r'^([\u4e00-\u9fa5]+)\d+[+]',  # 匹配 "算力8+" 或 "医疗医药12+"
            r'^([\u4e00-\u9fa5]+)[\*×]$',  # 匹配 "光通信*" 或 "新能源×"
            r'^([\u4e00-\u9fa5]{2,10})[^\u4e00-\u9fa5]',  # 匹配2-10个中文开头的
            r'^([\u4e00-\u9fa5]{2,10})$',  # 匹配纯中文题材名
        ]
        
        for pattern in patterns:
            match = re.search(pattern, line)
            if match:
                theme = match.group(1)
                # 过滤掉非题材的行
                if self._is_valid_theme(theme):
                    return theme
        
        return None
    
    def _is_valid_theme(self, text):
        """
        判断是否是有效的题材名称
        """
        # 排除常见的非题材词汇
        invalid_keywords = ['板数', '代码', '个股', '涨停', '时间', '流通', '市值', 
                           '成交额', '关键词', '不含', '统计', '数据', '根据', '市场',
                           '信息', '综合', '人工', '编写', '全网', '查看', '详细']
        
        for keyword in invalid_keywords:
            if keyword in text:
                return False
        
        return True
    
    def _extract_stock_line(self, row):
        """
        从行中提取股票信息
        :param row: 行数据（元组）
        :return: 股票信息字典
        """
        # 按照Excel列结构提取数据
        # A列（索引0）：板数
        # B列（索引1）：代码
        # C列（索引2）：个股名称
        # D列（索引3）：涨停时间
        # G列（索引6）：涨停关键词
        
        code = None
        name = None
        board_count = 1  # 默认首板
        time = None  # 涨停时间
        reason = None  # 涨停关键词
        
        # 提取代码（B列，索引1）
        if len(row) > 1 and row[1]:
            cell_value = str(row[1])
            code_match = re.search(r'(\d{6})', cell_value)
            if code_match:
                code = code_match.group(1)
        
        # 提取名称（C列，索引2）
        if len(row) > 2 and row[2]:
            name = str(row[2])
        
        # 提取涨停时间（D列，索引3）
        if len(row) > 3 and row[3]:
            time_value = str(row[3])
            # 匹配时间格式，如 "10:02:21"
            time_match = re.search(r'(\d{2}:\d{2}:\d{2})', time_value)
            if time_match:
                time = time_match.group(1)
        
        # 提取涨停关键词（G列，索引6）
        if len(row) > 6 and row[6]:
            reason = str(row[6])
        
        # 提取板数（A列，索引0）
        if len(row) > 0 and row[0]:
            board_value = str(row[0])
            # 匹配板数，如 "9天7板"、"2板" 或纯数字
            board_match = re.search(r'(\d+)板', board_value)
            if board_match:
                board_count = int(board_match.group(1))
            elif board_value.strip().isdigit():
                num = int(board_value.strip())
                if 1 <= num <= 20:  # 合理的板数范围
                    board_count = num
        
        if code:
            # 根据股票代码判断市场
            market = self._determine_market(code)
            # 如果找不到名称，使用代码作为名称
            if not name:
                name = code
            
            return {
                'code': code,
                'market': market,
                'name': name,
                'board_count': board_count,
                'time': time,
                'reason': reason
            }
        
        return None
    
    def _determine_market(self, code):
        """
        根据股票代码判断市场
        :param code: 股票代码
        :return: 市场标识
        """
        if code.startswith('6'):
            return 'SH'  # 沪市
        elif code.startswith('0'):
            return 'SZ'  # 深市
        elif code.startswith('3'):
            return 'SZ'  # 创业板
        elif code.startswith('9'):
            return 'BJ'  # 北交所
        else:
            return 'SH'  # 默认沪市